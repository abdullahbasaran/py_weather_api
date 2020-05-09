import requests
import pprint
import pandas as pd
from unidecode import unidecode
import schedule
import time
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
from datetime import date, timedelta
import datetime
from openpyxl import load_workbook

# excel den dataframe e donusturme
excel_data_df = pd.read_excel('ilce_merkez_coord.xlsx')

for col in ['ILCEADI']:
    excel_data_df[col] = excel_data_df[col].apply(unidecode)  # turkce karakterler decode edildi

print(excel_data_df) #turkce karakterlerden bagimsiz terminale respose yazilmasi

i=1
# edate = date(2008, 5, 5)   # end date

dt = date(2020, 3, 1)
end = date.today()
step = datetime.timedelta(days=1)

'''
book = load_workbook('New_Excel.xlsx')
writer = pd.ExcelWriter('New_Excel.xlsx', engine='openpyxl')
writer.book = book
'''

while dt < end:
     
    sdate = str(dt.strftime('%Y-%m-%d'))
    
    ilce = excel_data_df.loc[0][5]
    kod_ilce = excel_data_df.loc[0][3]
    url = 'https://api.weatherstack.com/historical?access_key=51d7bfab86b02431a6a6cc5d489ea60b&historical_date='+sdate+'&query='+ilce
    res = requests.get(url)
    data = res.json()
    # time = data['location']['localtime'] #istenilen degerin cekilmesi
    temp = data['current']['temperature']  #istenilen degerin cekilmesi
    sunrise = data['historical'][sdate]['astro']['sunrise']
    sunset = data['historical'][sdate]['astro']['sunset']
    moonrise = data['historical'][sdate]['astro']['moonrise']
    moonset = data['historical'][sdate]['astro']['moonset']
    moon_illumination = data['historical'][sdate]['astro']['moon_illumination']
    #time = data['historical'][sdate]['astro']['hourly'][]

    '''
    print('Time : ',sdate)
    print('Temprature : ',temp)
    print('Sunrise : ',sunrise)
    print('Sunset : ',sunset)
    print('Moonrise : ',moonrise)
    print('Moonset : ',moonset)
    '''
    
    def WriteToExcel():
        df = pd.DataFrame({'ILCEKOD': [kod_ilce], 'TIME': [sdate], 'TEMPRATURE': [temp], 'SUNRISE': [sunrise], 'SUNSET': [
                          sunset], 'MOONRISE': [moonrise], 'MOONSET': [moonset]})  # excel tablosunda her bir sutun icin degerlerin  girilmesi
        writer = pd.ExcelWriter('New_Excel.xlsx')
        df.to_excel(writer,'Sheet1',index=False)
            #writer.save()
        # try to open an existing workbook
        writer.book = load_workbook('New_Excel.xlsx')
        
        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # read existing file
        
        reader = pd.read_excel(r'New_Excel.xlsx')
        # write out the new sheet
        df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
        writer.close()

    WriteToExcel()
    dt += step
    i=i+1
''' 


writer = ExcelWriter('Excel2.xlsx',mode='a')
df.to_excel(writer,'Sheet1',index=False)
#max = max_row
#max=+1
writer.save()

writer.sheets=dict((ws.title,ws) for ws in book.worksheets)
df.to_excel(writer,'Sheet1',startrow=i, index=False)

max = ws.max_row
for row, entry in enumerate(data1,start=1):
st.cell(row=row+max, column=1, value=entry)

writer.save()
'''

    


# schedule.every(10).seconds.do(ExeceleYaz) #excele sche    dule ile yazma
# while True:
#    schedule.run_pending()
    # time.sleep(10)


# print(data)

